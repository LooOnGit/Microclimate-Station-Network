import secrets
import openpyxl
import qrcode

def generate_secure_numeric_id(length):
    if length < 1:
        raise ValueError("Length must be at least 1.")

    # Chọn ngẫu nhiên các chữ số
    secure_numeric_id = ''.join(secrets.choice('0123456789') for _ in range(length))
    return secure_numeric_id


def qrCode(data):
    # Tạo mã QR
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(data)
    qr.make(fit=True)

    # Tạo hình ảnh mã QR
    qr_img = qr.make_image(fill_color="black", back_color="white")

    # Lưu hình ảnh mã QR
    qr_img.save("QRGate/"+str(data)+".png")
def writeFile(idNew):
    path = "gateID.xlsx"
    wbObj = openpyxl.load_workbook(path)
    sheetObj = wbObj.active
    maxRow = sheetObj.max_row
    print(maxRow)

    if maxRow == 1:
        qrCode(idNew)
        id = sheetObj.cell(row=2, column=1)
        id.value = str(idNew)
    else:
        #Check ID already exists
        temp = True
        for i in range(2, maxRow+1):
            cellObj = sheetObj.cell(row=i, column=1)
            print(cellObj.value)
            if cellObj.value == str(idNew):
                temp = False
        print(temp)
        #add ID new
        if temp == True:
            qrCode(idNew)
            cellObj = sheetObj.cell(row=maxRow + 1, column=1)
            cellObj.value = idNew

    wbObj.save(path)


secure_numeric_id = 'GW'+ generate_secure_numeric_id(4)

print("Secure Numeric ID:", secure_numeric_id)
writeFile(secure_numeric_id)